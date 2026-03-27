"""Выгрузка замеров в Excel (.xlsx)."""

from __future__ import annotations

# GRACE[M-SVC-EXPORT][DOMAIN][BLOCK_XlsxExport]
# CONTRACT: measurements_to_xlsx_bytes — листы замеров и ошибок.

from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.datetime_util import iso_utc_z
from app.models import Measurement, MonitoredTarget, Provider

_ALL_HEADERS = [
    "id",
    "Время (UTC, ISO)",
    "Провайдер",
    "Модель",
    "target_id",
    "batch_id",
    "run_index",
    "Успех",
    "Сообщение об ошибке",
    "HTTP",
    "TTFT, с",
    "Total, с",
    "prompt_tokens",
    "completion_tokens",
    "Символов ответа",
    "gen t/s",
    "e2e t/s",
    "Стрим",
    "Чанков",
    "usage из API",
    "Δ mean, с",
    "Δ max, с",
]


def _load_rows(db: Session, *, hours: int, target_id: int | None) -> list[Measurement]:
    since = datetime.now(timezone.utc) - timedelta(hours=hours)
    q = (
        select(Measurement)
        .where(Measurement.created_at >= since)
        .options(joinedload(Measurement.target).joinedload(MonitoredTarget.provider))
        .order_by(Measurement.created_at)
    )
    if target_id is not None:
        q = q.where(Measurement.target_id == target_id)
    return list(db.scalars(q).unique().all())


def _row_cells(m: Measurement) -> list[object]:
    t = m.target
    p = t.provider
    return [
        m.id,
        iso_utc_z(m.created_at),
        p.display_name,
        t.model_name,
        t.id,
        m.batch_id,
        m.run_index,
        "да" if m.success else "нет",
        m.error_message or "",
        m.http_status if m.http_status is not None else "",
        m.ttft_s,
        m.total_s,
        m.prompt_tokens if m.prompt_tokens is not None else "",
        m.completion_tokens if m.completion_tokens is not None else "",
        m.output_chars,
        m.gen_tps if m.gen_tps is not None else "",
        m.e2e_tps if m.e2e_tps is not None else "",
        "да" if m.stream else "нет",
        m.chunk_count,
        "да" if m.usage_from_api else "нет",
        m.inter_chunk_gap_mean_s if m.inter_chunk_gap_mean_s is not None else "",
        m.inter_chunk_gap_max_s if m.inter_chunk_gap_max_s is not None else "",
    ]


def _write_sheet(ws, rows: Sequence[Measurement]) -> None:
    bold = Font(bold=True)
    for col, h in enumerate(_ALL_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
    for r_idx, m in enumerate(rows, start=2):
        for c_idx, val in enumerate(_row_cells(m), start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)


def measurements_to_xlsx_bytes(db: Session, *, hours: int, target_id: int | None) -> bytes:
    rows = _load_rows(db, hours=hours, target_id=target_id)
    err_rows = [m for m in rows if not m.success]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Замеры"
    _write_sheet(ws_all, rows)

    ws_err = wb.create_sheet("Ошибки")
    _write_sheet(ws_err, err_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
